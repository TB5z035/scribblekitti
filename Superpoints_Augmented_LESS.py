import numpy as np
import yaml
import os
import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from src.transforms.neighbors import knn_1
import numpy as np
import torch
from src.utils import sizes_to_pointers
from pgeof import pgeof
from src.data import Data
from sklearn.linear_model import RANSACRegressor
from src.transforms.partition import CutPursuitPartition, GridPartition
import matplotlib.pyplot as plt
from src.transforms import ConnectIsolated

class LESS():
    def __init__(self,config,config_LESS,args,split='train', label_directory='scribbles',target_directory='LESS') -> None:
        self.config = config
        self.split = split
        self.label_directory = label_directory
        self.true_directory = 'labels'
        self.fusion_target_directory = config_LESS['fusion_target_directory']
        self.target_directory = config_LESS['target_directory']
        self.target_directory_2 = config_LESS['target_directory_2']
        self.target_directory_3 = config_LESS['target_directory_3']
        self.target_directory_4 = config_LESS['target_directory_4']
        self.t = config_LESS['RANSAC']['scans_per_subsequence']
        self.l_grid = config_LESS['RANSAC']['l_grid']
        self.dist = config_LESS['RANSAC']['dist']
        self.iter_max = config_LESS['RANSAC']['iter_max']
        self.percent = config_LESS['RANSAC']['percent']
        self.d = config_LESS['cluster']['d']
        self.tic = 0
        self.Save = config_LESS['save']
        self.OOM_bar = config_LESS['RANSAC']['OOM_bar']
        self.save_file_number = config_LESS['save_file_number']
        self.seq_to_process = int(args.solve_seq)
        self.start_number_in_one_seq = config_LESS['start_number_in_one_seq']
        self.process_len_in_on_seq = config_LESS['process_len_in_on_seq']
        self.print_detail = config_LESS['print_detail']
        self.add_list = config_LESS['add_list']
        self.z_sort_bar = config_LESS['RANSAC']['z_sort_bar']
        self.count_classify_error = np.zeros([1,20])
        self.class_name_map = config['labels']

    def map_label(self,label, map_dict):
        maxkey = 0
        for key, data in map_dict.items():
            if isinstance(data, list):
                nel = len(data)
            else:
                nel = 1
            if key > maxkey:
                maxkey = key
        if nel > 1:
            lut = np.zeros((maxkey + 100, nel), dtype=np.int32)
        else:
            lut = np.zeros((maxkey + 100), dtype=np.int32)
        for key, data in map_dict.items():
            try:
                lut[key] = data
            except IndexError:
                print("Wrong key ", key)
        return lut[label]

    def run_LESS(self):
        root_dir = self.config['root_dir']
        cnt_seq = 0
        Calculate_Result_all = np.zeros([19,5],dtype=np.int64)      # row represents class_name, column is [Weak_Wrong, Propogated_Wrong, Weak_Num, Propogated_Num, All_Point]
        self.save_dir = './' +'LESS_Calculate/' + self.fusion_target_directory
        if not os.path.exists(self.save_dir):
            os.makedirs(self.save_dir)
        for seq in self.config['split'][self.split]:
            if self.seq_to_process >= 0:
                if seq != self.seq_to_process:
                    continue
            seq = '{0:02d}'.format(int(seq))
            self.label_class_dir = os.path.join(root_dir, seq,self.fusion_target_directory, 'group')
            self.labels_dir = os.path.join(root_dir, seq,self.fusion_target_directory, 'labels')
            if not os.path.exists(self.label_class_dir):
                    os.makedirs(self.label_class_dir)
            if not os.path.exists(self.labels_dir):
                    os.makedirs(self.labels_dir)   
            
            Calculate_Result = np.zeros([19,5],dtype=np.int64)      # row represents class_name, column is [Weak_Wrong, Propogated_Wrong, Weak_Num, Propogated_Num, All_Point]


            lidar_dir = os.path.join(root_dir, seq, 'velodyne')
            lidar_path = [os.path.join(dp, f) for dp, dn, fn in os.walk(os.path.expanduser(lidar_dir)) for f in fn if f.endswith('.bin')]
            self.seq_len = len(lidar_path)
            self.scans_cnt = 0
            label_dir = os.path.join(root_dir, seq, self.label_directory)
            true_label_dir = os.path.join(root_dir,seq,self.true_directory)
            label_path = [os.path.join(dp, f) for dp, dn, fn in os.walk(os.path.expanduser(label_dir)) for f in fn if f.endswith('.label')]
            true_label_path = [os.path.join(dp, f) for dp, dn, fn in os.walk(os.path.expanduser(true_label_dir)) for f in fn if f.endswith('.label')]
            # assert (len(lidar_path) == len(label_path))
            # label_paths.extend(label_path)
            LESS_label_dir = os.path.join(root_dir, seq, self.target_directory,'labels')
            LESS_labels_path = [os.path.join(dp, f) for dp, dn, fn in os.walk(os.path.expanduser(LESS_label_dir)) for f in fn if f.endswith('.label')]
            # assert (len(lidar_paths) == len(LESS_labels_paths))
            label_group_dir = os.path.join(root_dir, seq, self.target_directory,'group')
            label_group_paths = [os.path.join(dp, f) for dp, dn, fn in os.walk(os.path.expanduser(label_group_dir)) for f in fn if f.endswith('.label')]

            LESS_label_dir_2 = os.path.join(root_dir, seq, self.target_directory_2,'labels')
            LESS_labels_path_2 = [os.path.join(dp, f) for dp, dn, fn in os.walk(os.path.expanduser(LESS_label_dir_2)) for f in fn if f.endswith('.label')]
            # assert (len(lidar_paths) == len(LESS_labels_paths))
            label_group_dir_2 = os.path.join(root_dir, seq, self.target_directory_2,'group')
            label_group_paths_2 = [os.path.join(dp, f) for dp, dn, fn in os.walk(os.path.expanduser(label_group_dir_2)) for f in fn if f.endswith('.label')]

            LESS_label_dir_3 = os.path.join(root_dir, seq, self.target_directory_3,'labels')
            LESS_labels_path_3 = [os.path.join(dp, f) for dp, dn, fn in os.walk(os.path.expanduser(LESS_label_dir_3)) for f in fn if f.endswith('.label')]
            # assert (len(lidar_paths) == len(LESS_labels_paths))
            label_group_dir_3 = os.path.join(root_dir, seq, self.target_directory_3,'group')
            label_group_paths_3 = [os.path.join(dp, f) for dp, dn, fn in os.walk(os.path.expanduser(label_group_dir_3)) for f in fn if f.endswith('.label')]

            LESS_label_dir_4 = os.path.join(root_dir, seq, self.target_directory_4,'labels')
            LESS_labels_path_4 = [os.path.join(dp, f) for dp, dn, fn in os.walk(os.path.expanduser(LESS_label_dir_4)) for f in fn if f.endswith('.label')]
            # assert (len(lidar_paths) == len(LESS_labels_paths))
            label_group_dir_4 = os.path.join(root_dir, seq, self.target_directory_4,'group')
            label_group_paths_4 = [os.path.join(dp, f) for dp, dn, fn in os.walk(os.path.expanduser(label_group_dir_4)) for f in fn if f.endswith('.label')]

            pose_path = os.path.join(root_dir,seq,'poses.txt')
            self.fusion(lidar_path=lidar_path,pose_path=pose_path,label_path=label_path,t=self.t,l_grid=self.l_grid,dist=self.dist,iter_max=self.iter_max,percent=self.percent,
                        d=self.d,seq=seq,true_label_path=true_label_path,LESS_labels_path=LESS_labels_path,label_group_paths=label_group_paths,Calculate_Result=Calculate_Result,Calculate_Result_all=Calculate_Result_all,
                        LESS_labels_path_2=LESS_labels_path_2,label_group_paths_2=label_group_paths_2,
                        LESS_labels_path_3=LESS_labels_path_3,label_group_paths_3=label_group_paths_3,
                        LESS_labels_path_4=LESS_labels_path_4,label_group_paths_4=label_group_paths_4
                        )
        Result_list = []
        for i in range(1,20):
            class_name = self.class_name_map[i]
            if (Calculate_Result_all[i-1,2]!= 0) & (Calculate_Result_all[i-1,3]!=0):
                one_line_list = [class_name,Calculate_Result_all[i-1,0], Calculate_Result_all[i-1,1], Calculate_Result_all[i-1,2], Calculate_Result_all[i-1,3], Calculate_Result_all[i-1,4],Calculate_Result_all[i-1,0]/Calculate_Result_all[i-1,2],Calculate_Result_all[i-1,1]/Calculate_Result_all[i-1,3]]
            else:
                one_line_list = [class_name,0,0,0,0]
            Result_list.append(one_line_list)
        columns = ["Class Name", "Weak_Wrong", "Propogated_Wrong", "Weak_Num", "Propogated_Num", "Points_Num",'Weak_Wrong_Percent','Propogated_Wrong_Percent']
        df = pd.DataFrame(Result_list, columns=columns)
        df_sorted = df.sort_values(by="Points_Num", ascending=False)
        excel_filename = self.save_dir + '/' + self.fusion_target_directory +'_' + "all "+".xlsx"
        df_sorted.to_excel(excel_filename, index=False)
        # self.Preconfig_Excel(excel_filename=excel_filename,df=df_sorted)
        return 1
    
    def AddKeysTo_process_single_key(self, data, key, to, AddKeys_config):
        # Read existing features and the attribute of interest
        feat = getattr(data, key, None)
        x = getattr(data, to, None)

        # Skip if the attribute is None
        if feat is None:
            if AddKeys_config['strict']:
                raise Exception(f"Data should contain the attribute '{key}'")
            else:
                return data

        # Remove the attribute from the Data, if required
        if AddKeys_config['delete_after']:
            delattr(data, key)

        # In case Data has no features yet
        if x is None:
            if AddKeys_config['strict'] and data.num_nodes != feat.shape[0]:
                raise Exception(f"Data should contain the attribute '{to}'")
            if feat.dim() == 1:
                feat = feat.unsqueeze(-1)
            data[to] = feat
            return data

        # Make sure shapes match
        if x.shape[0] != feat.shape[0]:
            raise Exception(
                f"The tensors '{to}' and '{key}' can't be concatenated, "
                f"'{to}': {x.shape[0]}, '{key}': {feat.shape[0]}")

        # Concatenate x and feat
        if x.dim() == 1:
            x = x.unsqueeze(-1)
        if feat.dim() == 1:
            feat = feat.unsqueeze(-1)
        data[to] = torch.cat([x, feat], dim=-1)

        return data

    def AddKeyTo_process(self, data, AddKeys_config):
        if AddKeys_config['keys'] is None or len(AddKeys_config['keys']) == 0:
            return data

        for key in AddKeys_config['keys']:
            data = self.AddKeysTo_process_single_key(data, key, AddKeys_config['to'],AddKeys_config)

        return data


    def get_lidar(self, path):
        lidar = np.fromfile(path, dtype=np.float32)
        lidar = torch.from_numpy(lidar.reshape((-1, 4)))
        return lidar[:,:3]


    def GroundElevation_process(self, data, GroundElevation_config):
        # Recover the point positions
        pos = data.pos.cpu().numpy()

        # To avoid capturing high above-ground flat structures, we only
        # keep points which are within `threshold` of the lowest point.
        idx_low = np.where(pos[:, 2] - pos[:, 2].min() < GroundElevation_config['threshold'])[0]

        # Search the ground plane using RANSAC
        ransac = RANSACRegressor(min_samples=0.1,random_state=0, residual_threshold=1e-3).fit(pos[idx_low, :2], pos[idx_low, 2])

        # Compute the pointwise elevation as the distance to the plane
        # and scale it
        h = pos[:, 2] - ransac.predict(pos[:, :2])
        h = h / GroundElevation_config['scale']

        # Save in Data attribute `elevation`
        data.elevation = torch.from_numpy(h).to(data.device).view(-1, 1)

        return data

    def Preconfig_Excel(self,excel_filename,df):
        wb = load_workbook(excel_filename)
        ws = wb.active
        # Center align all cells
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Format "Weak_Num" and "Propogated_Num" columns as percentages
        for col_letter in ["Weak_Wrong_Percent", "Propogated_Wrong_Percent"]:
            for row in ws.iter_rows(min_row=2, min_col=df.columns.get_loc(col_letter) + 1, max_row=ws.max_row, max_col=df.columns.get_loc(col_letter) + 1):
                for cell in row:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00

        # Adjust column widths based on content
        for column in df.columns:
            column_idx = df.columns.get_loc(column) + 1
            max_length = max(df[column].astype(str).apply(len).max(), len(column))
            adjusted_width = (max_length + 2) * 1.2  # Adding a little extra space
            ws.column_dimensions[get_column_letter(column_idx)].width = adjusted_width
        wb.save(excel_filename)

    # t represent t scans to be combined with
    # l_grid means we use l*l grid to run RANSAC
    def fusion(self,lidar_path,pose_path,label_path,t,l_grid,dist,iter_max,percent,d,seq,true_label_path,LESS_labels_path,label_group_paths,Calculate_Result,Calculate_Result_all,LESS_labels_path_2,LESS_labels_path_3,label_group_paths_2,label_group_paths_3,LESS_labels_path_4,label_group_paths_4):     
        lidar_path.sort()
        label_path.sort()
        true_label_path.sort()
        LESS_labels_path.sort()
        label_group_paths.sort()
        LESS_labels_path_2.sort()
        label_group_paths_2.sort()
        LESS_labels_path_3.sort()
        label_group_paths_3.sort()
        LESS_labels_path_4.sort()
        label_group_paths_4.sort()
        cnt = 0

        # for lidar_file,label_file,true_label_file,LESS_label_file,label_group_file,LESS_label_file_2,label_group_file_2,LESS_label_file_3,label_group_file_3,LESS_label_file_4,label_group_file_4 in zip(lidar_path,label_path,true_label_path,LESS_labels_path,label_group_paths,LESS_labels_path_2,label_group_paths_2,LESS_labels_path_3,label_group_paths_3,LESS_labels_path_4,label_group_paths_4):
        for lidar_file,label_file,true_label_file,LESS_label_file,label_group_file,LESS_label_file_3,label_group_file_3,LESS_label_file_4,label_group_file_4 in zip(lidar_path,label_path,true_label_path,LESS_labels_path,label_group_paths,LESS_labels_path_3,label_group_paths_3,LESS_labels_path_4,label_group_paths_4):
            if self.start_number_in_one_seq != 0:
                if cnt < self.start_number_in_one_seq:
                    print('pass seq',seq,'scan',cnt)
                    cnt += 1
                    continue
                elif cnt >= self.start_number_in_one_seq + self.process_len_in_on_seq:
                    print('seq',seq,'process',self.start_number_in_one_seq,'to',self.start_number_in_one_seq+self.process_len_in_on_seq,'finish')
                    return
            # start a new subsequence 
            lidar_xyz_unified = np.array([[0,0,0]],dtype=np.float32)
            label_unified = np.array([[0]])
            true_label_unified = np.array([[0]])

            # read the pose and aggregate scans
            # pose_SE3 = np.array([[r11,r12,r13,t14],[r21,r22,r23,t24],[r31,r32,r33,t34],[0,0,0,1]])
            # pose_SE3_inv = np.linalg.inv(pose_SE3)
            # (r11,r12,r13,t14),(r21,r22,r23,t24),(r31,r32,r33,t34),(_,_,_,_)=pose_SE3_inv
            lidar_data = np.fromfile(lidar_file,dtype=np.float32)
            lidar_data = lidar_data.reshape((-1,4))
            lidar_x = lidar_data[:,0]
            lidar_y = lidar_data[:,1]
            lidar_z = lidar_data[:,2]
            label = np.fromfile(label_file, dtype=np.int32)
            label = label.reshape((-1)) & 0xFFFF
            label = self.map_label(label, self.config['learning_map'])
            true_label = np.fromfile(true_label_file, dtype=np.int32)
            true_label = true_label.reshape((-1)) & 0xFFFF
            true_label = self.map_label(true_label, self.config['learning_map'])
            # Matrix_RT = np.array([r11,r12,r13,t11],[r21,r22,r23,t12],[r31,r32,r33,t13],[0,0,0,1])
            LESS_label = np.fromfile(LESS_label_file, dtype=bool)
            LESS_label = LESS_label.reshape((-1, 20))
            label_group = np.fromfile(label_group_file,dtype=np.int8)

            # LESS_label_2 = np.fromfile(LESS_label_file_2, dtype=bool)
            # LESS_label_2 = LESS_label_2.reshape((-1, 20))
            # label_group_2 = np.fromfile(label_group_file_2,dtype=np.int8)

            # LESS_label_3 = np.fromfile(LESS_label_file_3, dtype=bool)
            # LESS_label_3 = LESS_label_3.reshape((-1, 20))
            # label_group_3 = np.fromfile(label_group_file_3,dtype=np.int8)


            # LESS_label_4 = np.fromfile(LESS_label_file_4, dtype=bool)
            # LESS_label_4 = LESS_label_4.reshape((-1, 20))
            # label_group_4 = np.fromfile(label_group_file_4,dtype=np.int8)
            # lidar_x = lidar_x*r11 +r12*lidar_y+r13*lidar_z + t14
            # lidar_y = lidar_x*r21 +r22*lidar_y+r23*lidar_z + t24
            # lidar_z = lidar_x*r31 +r32*lidar_y+r33*lidar_z + t34

            # label_group_2_selected = label_group_2[nothing_index]         # select the noting group in main LESS result index in LESS result 2
            # label_group_3_selected = label_group_3[nothing_index]         # select the noting group in main LESS result index in LESS result 3

            # LESS_label[nothing_index] += LESS_label_2[nothing_index] + LESS_label_3[nothing_index]
            # label_group[nothing_index] = 3
            file_name = LESS_label_file.split('/labels/')[1]
            label_class_file = os.path.join(self.label_class_dir,file_name)
            labels_file = os.path.join(self.labels_dir,file_name)

######################################################## Get_SuperPoints_Index ########################################################
            data = Data()
            data.pos = torch.from_numpy(lidar_data[:,:3])
            # data = torch.from_numpy(data)
            config = {'keys':['linearity','planarity','scattering','verticality','curvature','length','surface','volume','normal'],'AdjacencyGraph':{'k':10,'w':1}}

            ##################################################### KNN #####################################################
            neighbors, distances = knn_1(data.pos[:,:3], k=45, r_max=2, oversample=False,self_is_neighbor=False, verbose=False)
            # neighbors, distances = knn_1(data[:,:3], k=config['knn']['k'], r_max=config['knn']['r_max'], oversample=False,self_is_neighbor=False, verbose=False)
            data.neighbor_index = neighbors
            data.neighbor_distance = distances


            ##################################################### GroundElevation #####################################################
            GroundElevation_config = {'threshold': 5,'scale':20}
            data = self.GroundElevation_process(data,GroundElevation_config)


            ##################################################### Point-Features #####################################################
            assert data.has_neighbors, \
                        "Data is expected to have a 'neighbor_index' attribute"
            assert data.num_nodes < np.iinfo(np.uint32).max, \
                        "Too many nodes for `uint32` indices"
            assert data.neighbor_index.max() < np.iinfo(np.uint32).max, \
                        "Too high 'neighbor_index' indices for `uint32` indices"
            # get point-wise density feature
            dmax = data.neighbor_distance.max(dim=1).values
            k = data.neighbor_index.ge(0).sum(dim=1)
            data.density = (k / dmax ** 2).view(-1, 1)


            # get geometric features 
            # device = data.pos.device
            # xyz = data[:,:3].cpu().numpy()
            device = data.pos.device
            xyz = data.pos.cpu().numpy()
            nn = torch.cat(
                (torch.arange(xyz.shape[0]).view(-1, 1), data.neighbor_index),
                dim=1)
            k = nn.shape[1]

            # Check for missing neighbors (indicated by -1 indices)
            n_missing = (nn < 0).sum(dim=1)
            if (n_missing > 0).any():
                sizes = k - n_missing
                nn = nn[nn >= 0]
                nn_ptr = sizes_to_pointers(sizes.cpu())
            else:
                nn = nn.flatten().cpu()
                nn_ptr = torch.arange(xyz.shape[0] + 1) * k
            nn = nn.numpy().astype('uint32')
            nn_ptr = nn_ptr.numpy().astype('uint32')

            # Make sure array are contiguous before moving to C++
            xyz = np.ascontiguousarray(xyz)
            nn = np.ascontiguousarray(nn)
            nn_ptr = np.ascontiguousarray(nn_ptr)
            # C++ geometric features computation on CPU
            # f = pgeof(
            #     xyz, nn, nn_ptr, k_min=config['pgeof']['k_min'], k_step=config['pgeof']['k_step'],
            #     k_min_search=config['pgeof']['k_min_search'], verbose=False)
            f = pgeof(
                xyz, nn, nn_ptr, k_min=1, k_step=-1,
                k_min_search=25, verbose=False)
            f = torch.from_numpy(f.astype('float32'))
            if 'linearity' in config['keys']:
                # data.linearity = f[:, 0].view(-1, 1).to(device)
                data.linearity = f[:, 0].view(-1, 1)

            if 'planarity' in config['keys']:
                # data.planarity = f[:, 1].view(-1, 1).to(device)
                data.planarity = f[:, 1].view(-1, 1)

            if 'scattering' in config['keys']:
                # data.scattering = f[:, 2].view(-1, 1).to(device)
                data.scattering = f[:, 2].view(-1, 1)

            # Heuristic to increase importance of verticality in
            # partition
            if 'verticality' in config['keys']:
                # data.verticality = f[:, 3].view(-1, 1).to(device)
                data.verticality = f[:, 3].view(-1, 1)
                data.verticality *= 2

            if 'curvature' in config['keys']:
                # data.curvature = f[:, 10].view(-1, 1).to(device)
                data.curvature = f[:, 10].view(-1, 1)

            if 'length' in config['keys']:
                # data.length = f[:, 7].view(-1, 1).to(device)
                data.length = f[:, 7].view(-1, 1)

            if 'surface' in config['keys']:
                # data.surface = f[:, 8].view(-1, 1).to(device)
                data.surface = f[:, 8].view(-1, 1)

            if 'volume' in config['keys']:
                # data.volume = f[:, 9].view(-1, 1).to(device)
                data.volume = f[:, 9].view(-1, 1)

            # As a way to "stabilize" the normals' orientation, we
            # choose to express them as oriented in the z+ half-space
            if 'normal' in config['keys']:
                # data.normal = f[:, 4:7].view(-1, 3).to(device)
                data.normal = f[:, 4:7].view(-1, 3)
                data.normal[data.normal[:, 2] < 0] *= -1


            ##################################################### AdjacencyGraph #####################################################

            AdjacencyGraph_k = config['AdjacencyGraph']['k']
            AdjacencyGraph_w = config['AdjacencyGraph']['w']

            # Compute source and target indices based on neighbors
            source = torch.arange(
                data.num_nodes, device=data.device).repeat_interleave(AdjacencyGraph_k)
            target = data.neighbor_index[:, :AdjacencyGraph_k].flatten()

            # Account for -1 neighbors and delete corresponding edges
            mask = target >= 0
            source = source[mask]
            target = target[mask]

            # Save edges and edge features in data
            data.edge_index = torch.stack((source, target))
            if AdjacencyGraph_w > 0:
                # Recover the neighbor distances and apply the masking
                distances_After_Ad = distances[:, :AdjacencyGraph_k].flatten()[mask]
                data.edge_attr = 1 / (AdjacencyGraph_w + distances_After_Ad / distances_After_Ad.mean())
            else:
                data.edge_attr = torch.ones_like(source, dtype=torch.float)

            # ConnectIsolated
            ConnectIsolated_config = {'k':1}
            ConnectIsolated_Transforms = ConnectIsolated(k=ConnectIsolated_config['k'])

            # AddKeysTo
            AddKeys_config = {'keys':['linearity','planarity','scattering','verticality','elevation'],'to':'x','delete_after':False,'strict':True}
            data = self.AddKeyTo_process(data=data, AddKeys_config=AddKeys_config)

            # CutPursuitPartition
            CutPursuitPartition_config = {
                'pcp_regularization': [0.1, 0.2, 0.6],
                'pcp_spatial_weight': [1, 1e-1, 1e-2],
                'pcp_cutoff': [10, 30, 100],
                'pcp_k_adjacency': 10,
                'pcp_w_adjacency': 1,
                'pcp_iterations': 15,
                'parallel': True,
                'verbose': False}
            CutPursuitPartition_Transforms = CutPursuitPartition(regularization=CutPursuitPartition_config['pcp_regularization'],
                                                                spatial_weight=CutPursuitPartition_config['pcp_spatial_weight'],
                                                                cutoff = CutPursuitPartition_config['pcp_cutoff'],
                                                                parallel=CutPursuitPartition_config['parallel'],
                                                                iterations=CutPursuitPartition_config['pcp_iterations'],
                                                                k_adjacency=CutPursuitPartition_config['pcp_k_adjacency'],
                                                                verbose=CutPursuitPartition_config['verbose'])
            nag = CutPursuitPartition_Transforms(data)
            # points_with_rgb = get_points_with_rgb(data.pos,super_index_0)
            # save_matrix_to_txt(points_with_rgb,'level_1_super_index_points.txt')
            # points_with_rgb = get_points_with_rgb(data.pos,(nag[1].super_index)[nag[0].super_index])
            # save_matrix_to_txt(points_with_rgb,'level_2_super_index_points.txt')
            # points_with_rgb = get_points_with_rgb(data.pos,nag[2].super_index[(nag[1].super_index)[nag[0].super_index]])
            # save_matrix_to_txt(points_with_rgb,'level_3_super_index_points.txt')
            # points_with_rgb = get_points_with_rgb(data.pos,nag[3].super_index[nag[2].super_index[(nag[1].super_index)[nag[0].super_index]]])
            # save_matrix_to_txt(points_with_rgb,'level_4_super_index_points.txt')
            
            # group_level_3 = nag[2].super_index[(nag[1].super_index)[nag[0].super_index]]
            # group_level_3 = (nag[1].super_index)[nag[0].super_index]
            group_level_3 = nag[0].super_index
            group_level_3 = np.array(group_level_3)
            scribble_index = np.where(label_group == 1)[0]
            group_level_3_in_scribble_index = group_level_3[scribble_index]
            scribble_label = label[scribble_index]
            LESS_label[scribble_index,:]
            for i in np.unique(group_level_3_in_scribble_index):
                label_save = np.zeros((1,20),dtype=bool)
                label_to_propogate = scribble_label[group_level_3_in_scribble_index == i]
                label_to_propogate_unique = np.unique(label_to_propogate)
                label_to_propogate_unique = label_to_propogate_unique[label_to_propogate_unique!=0]
                mask_to_propogated = group_level_3==i
                if label_to_propogate_unique.shape[0] == 1:
                    label_save[:,label_to_propogate_unique] = True
                    label_group[mask_to_propogated] = 2
                    LESS_label[mask_to_propogated] = label_save
                else:
                    label_save[:,label_to_propogate_unique] = True
                    label_group[mask_to_propogated] = 3
                    LESS_label[mask_to_propogated] = label_save
            # noting, scribbles, propogated, weak
            #   0         1          2         3 



#######################################################################################################################################
            # if seq == '06':
            #     if cnt == 396:
            #         a = 1
            # if seq != '06':
            #     nothing_index = np.where(label_group == 0)[0]
            #     nothing_index_in_origin_point_while_propogated_in_another = nothing_index[np.where(label_group_2[nothing_index]==2)]
            #     label_group[nothing_index_in_origin_point_while_propogated_in_another] = 2
            #     LESS_label[nothing_index_in_origin_point_while_propogated_in_another] = LESS_label_2[nothing_index_in_origin_point_while_propogated_in_another]

            # nothing_index = np.where(label_group == 0)[0]
            # nothing_index_in_origin_point_while_propogated_in_another = nothing_index[np.where(label_group_3[nothing_index]==2)]
            # label_group[nothing_index_in_origin_point_while_propogated_in_another] = 2
            # LESS_label[nothing_index_in_origin_point_while_propogated_in_another] = LESS_label_3[nothing_index_in_origin_point_while_propogated_in_another]

            # nothing_index = np.where(label_group == 0)[0]
            # nothing_index_in_origin_point_while_propogated_in_another = nothing_index[np.where(label_group_4[nothing_index]==2)]
            # label_group[nothing_index_in_origin_point_while_propogated_in_another] = 2
            # LESS_label[nothing_index_in_origin_point_while_propogated_in_another] = LESS_label_4[nothing_index_in_origin_point_while_propogated_in_another]

            
            # label_group_2[nothing_index][label_group_2[nothing_index] !=0].shape
            # label_group_3[nothing_index][label_group_3[nothing_index] !=0].shape

            lidar_x = np.expand_dims(lidar_x,axis=1)
            lidar_y = np.expand_dims(lidar_y,axis=1)
            lidar_z = np.expand_dims(lidar_z,axis=1)
            label   = np.expand_dims(label,  axis=1)
            lidar_xyz = np.concatenate((lidar_x,lidar_y,lidar_z),axis=1)
            lidar_xyz_unified = np.concatenate((lidar_xyz_unified,lidar_xyz),axis=0)        # OK
            label_unified = np.concatenate((label_unified,label),axis=0)        # OK
            true_label_unified = np.concatenate((true_label_unified,np.expand_dims(true_label,axis=1)),axis=0)        # OK
            lidar_xyz_unified = lidar_xyz_unified[1:,:]     # get all points in one axis
            label_unified = label_unified[1:,:]
            true_label_unified = true_label_unified[1:,:]


            propogated_label_true = true_label_unified[label_group==2]
            propogated_label = np.where(LESS_label[np.where(label_group == 2)[0],:] == True)[1]
            propogated_label = propogated_label[propogated_label!=0]
            True_propogated_point = np.count_nonzero(np.expand_dims(propogated_label,axis=1) == propogated_label_true)   
            print('seq ',seq,'scan ',cnt,'propogated points right:',True_propogated_point,'/',propogated_label.shape[0],'. Wrong is',(1 - True_propogated_point/propogated_label.shape[0])*100,'%',flush=True)
            weak_label_true = true_label_unified[label_group==3,:]
            # weak_label_true = np.expand_dims(weak_label_true,axis=1)
            weak_label = LESS_label[np.where(label_group == 3)[0],:] == True
            weak_label_right = weak_label[np.arange(len(weak_label_true)), weak_label_true.ravel()].reshape(-1, 1)
            print('seq ',seq,'scan ',cnt,'Weak points right:',weak_label_right.sum(),'/',weak_label.shape[0],'. Wrong is',(1 - weak_label_right.sum()/weak_label.shape[0])*100,'%',flush=True)

            label_group.tofile(label_class_file)
            LESS_label.tofile(labels_file)

            # row represents class_name, column is [Weak_Wrong, Propogated_Wrong, Weak_Num, Propogated_Num]
            for i in range(1,20):
                index = np.where(true_label_unified==i)[0]
                true_label_in_index = true_label_unified[index]
                LESS_label_in_index = LESS_label[index]
                label_group_in_index = label_group[index]
                propogated_index = np.where(label_group_in_index==2)[0]
                weak_inedx = np.where(label_group_in_index==3)[0]
                weak_label = LESS_label_in_index[weak_inedx]
                weak_label[:,0] = False
                true_label_in_weak_index = true_label_in_index[weak_inedx]
                weak_label_right = weak_label[np.arange(len(true_label_in_weak_index)), true_label_in_weak_index.ravel()].reshape(-1, 1)
                Calculate_Result[i-1,0] += weak_label_right.shape[0] - weak_label_right.sum()
                Calculate_Result[i-1,2] += weak_label_right.shape[0]
                Calculate_Result_all[i-1,0] += weak_label_right.shape[0] - weak_label_right.sum()
                Calculate_Result_all[i-1,2] += weak_label_right.shape[0]

                propogated_label = LESS_label_in_index[propogated_index]
                propogated_label = np.where(propogated_label==1)[1]
                propogated_label = propogated_label[propogated_label!=0]
                true_label_in_propogated_index = true_label_in_index[propogated_index]
                True_propogated_point = np.count_nonzero(np.expand_dims(propogated_label,axis=1) == true_label_in_propogated_index)
                Calculate_Result[i-1,1] += true_label_in_propogated_index.shape[0] - True_propogated_point
                Calculate_Result[i-1,3] += true_label_in_propogated_index.shape[0]        
                Calculate_Result[i-1,4] += index.shape[0]
                Calculate_Result_all[i-1,1] += true_label_in_propogated_index.shape[0] - True_propogated_point
                Calculate_Result_all[i-1,3] += true_label_in_propogated_index.shape[0]        
                Calculate_Result_all[i-1,4] += index.shape[0]
            cnt += 1
            
            print('seq ',seq,'scan ',cnt,' finish')
        Result_list = []
        for i in range(1,20):
            class_name = self.class_name_map[i]
            if (Calculate_Result[i-1,2]!= 0) & (Calculate_Result[i-1,3]!=0):
                one_line_list = [class_name,Calculate_Result[i-1,0], Calculate_Result[i-1,1], Calculate_Result[i-1,2], Calculate_Result[i-1,3], Calculate_Result[i-1,4],Calculate_Result[i-1,0]/Calculate_Result[i-1,2],Calculate_Result[i-1,1]/Calculate_Result[i-1,3]]
            else:
                one_line_list = [class_name,0,0,0,0]
            Result_list.append(one_line_list)
        columns = ["Class Name", "Weak_Wrong", "Propogated_Wrong", "Weak_Num", "Propogated_Num", "Points_Num",'Weak_Wrong_Percent','Propogated_Wrong_Percent']
        df = pd.DataFrame(Result_list, columns=columns)
        df_sorted = df.sort_values(by="Points_Num", ascending=False)
        excel_filename = self.save_dir +'/'+ self.fusion_target_directory +'_' + "seq "+seq+".xlsx"
        df_sorted.to_excel(excel_filename, index=False)
        # self.Preconfig_Excel(excel_filename=excel_filename,df=df_sorted)
        self.Preconfig_Excel(excel_filename=excel_filename,df=df_sorted)


if __name__ == '__main__':
    # config_path = 'config/LESS_dist3.yaml'
    config_path = 'config/LESS_fusion.yaml'
    dataset_config_path = 'config/dataset/semantickitti.yaml'

    with open(config_path, 'r') as f:
        config = yaml.safe_load(f)
    with open(dataset_config_path, 'r') as f:
        config['dataset'].update(yaml.safe_load(f))
    with open(dataset_config_path, 'r') as f:
        config['val_dataset'].update(yaml.safe_load(f))
    parser = argparse.ArgumentParser()
    parser.add_argument('--solve_seq', default=0,help='-1 means all,0-10 should be input')
    args = parser.parse_args()
    less = LESS(config['dataset'],config['LESS'],args=args)
    less.run_LESS()